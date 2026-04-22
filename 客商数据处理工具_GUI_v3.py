#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SAP MDG 客商数据处理工具 - GUI版本 v3
多公司混排 · 文件名解析 · 保留格式
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import sys
import shutil
import re

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("正在安装依赖...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl import load_workbook


class ModernSAPMDGTool:
    def __init__(self, root):
        self.root = root
        self.root.title("SAP MDG 客商数据处理工具 v3 - 多公司混排")
        self.root.geometry("900x800")
        self.root.minsize(700, 600)
        
        # 设置DPI感知（Windows高分屏支持）
        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except:
            pass
        
        # 配置ttk样式 - 现代风格
        self.setup_styles()
        
        # 数据变量
        self.data_files = []  # 多文件列表，每项: {'path': str, 'company': str, 'group': str}
        self.template_file = tk.StringVar()
        self.exclude_file = tk.StringVar()
        self.output_dir = tk.StringVar(value=os.path.expanduser("~/Desktop"))
        self.split_z012 = tk.StringVar(value="200")
        self.split_z001 = tk.StringVar(value="500")
        
        self.create_widgets()
    
    def setup_styles(self):
        """配置现代化样式"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # 配色方案 - 现代蓝紫色系
        self.colors = {
            'primary': '#6366f1',      # 主色：靛蓝
            'primary_dark': '#4f46e5', # 深色
            'secondary': '#8b5cf6',    # 紫色
            'bg': '#f8fafc',           # 背景
            'card': '#ffffff',         # 卡片
            'text': '#1e293b',         # 主文字
            'text_muted': '#64748b',   # 次要文字
            'border': '#e2e8f0',       # 边框
            'success': '#22c55e',      # 成功
            'warning': '#f59e0b',      # 警告
            'error': '#ef4444',        # 错误
        }
        
        # 全局样式
        style.configure('.',
            font=('Microsoft YaHei', 10),
            background=self.colors['bg']
        )
        
        # 标签框架样式（卡片效果）
        style.configure('Card.TLabelframe',
            background=self.colors['card'],
            borderwidth=1,
            relief='solid',
            bordercolor=self.colors['border']
        )
        style.configure('Card.TLabelframe.Label',
            background=self.colors['card'],
            foreground=self.colors['text'],
            font=('Microsoft YaHei', 11, 'bold'),
            padding=(10, 5)
        )
        
        # 按钮样式
        style.configure('Primary.TButton',
            background=self.colors['primary'],
            foreground='white',
            font=('Microsoft YaHei', 10, 'bold'),
            padding=(20, 10),
            borderwidth=0,
            relief='flat'
        )
        style.map('Primary.TButton',
            background=[('active', self.colors['primary_dark']), ('pressed', self.colors['primary_dark'])],
            foreground=[('active', 'white'), ('pressed', 'white')]
        )
        
        style.configure('Secondary.TButton',
            background='white',
            foreground=self.colors['text'],
            font=('Microsoft YaHei', 9),
            padding=(15, 6),
            borderwidth=1,
            relief='solid'
        )
        style.map('Secondary.TButton',
            background=[('active', self.colors['bg']), ('pressed', self.colors['border'])],
            bordercolor=[('active', self.colors['primary']), ('pressed', self.colors['primary'])]
        )
        
        # 输入框样式
        style.configure('Modern.TEntry',
            font=('Microsoft YaHei', 10),
            padding=8,
            borderwidth=1,
            relief='solid'
        )
        
        # 标签样式
        style.configure('Title.TLabel',
            font=('Microsoft YaHei', 20, 'bold'),
            foreground=self.colors['text'],
            background=self.colors['bg']
        )
        
        style.configure('Subtitle.TLabel',
            font=('Microsoft YaHei', 11),
            foreground=self.colors['text_muted'],
            background=self.colors['bg']
        )
        
        style.configure('Field.TLabel',
            font=('Microsoft YaHei', 10, 'bold'),
            foreground=self.colors['text'],
            background=self.colors['card']
        )
        
        style.configure('Hint.TLabel',
            font=('Microsoft YaHei', 9),
            foreground=self.colors['text_muted'],
            background=self.colors['card']
        )
        
        # 进度条样式
        style.configure('Horizontal.TProgressbar',
            thickness=6,
            background=self.colors['primary'],
            troughcolor=self.colors['border'],
            borderwidth=0,
            relief='flat'
        )
    
    def create_widgets(self):
        """创建主界面"""
        # 主容器 - 使用Canvas实现滚动
        self.main_canvas = tk.Canvas(self.root, bg=self.colors['bg'], highlightthickness=0)
        self.main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 滚动条
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.main_canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill="y")
        self.main_canvas.configure(yscrollcommand=scrollbar.set)
        
        # 内容框架
        self.content_frame = tk.Frame(self.main_canvas, bg=self.colors['bg'])
        self.canvas_window = self.main_canvas.create_window((0, 0), window=self.content_frame, anchor="nw")
        
        # 配置Canvas滚动
        self.content_frame.bind("<Configure>", self.on_frame_configure)
        self.main_canvas.bind("<Configure>", self.on_canvas_configure)
        
        # 绑定鼠标滚轮
        self.main_canvas.bind_all("<MouseWheel>", self.on_mousewheel)
        
        # 内边距
        padding = {'padx': 30, 'pady': 15}
        
        # ========== 头部区域 ==========
        header_frame = tk.Frame(self.content_frame, bg=self.colors['bg'])
        header_frame.pack(fill=tk.X, **padding)
        
        # 标题
        title = ttk.Label(header_frame, text="SAP MDG 客商数据处理工具 v3", style='Title.TLabel')
        title.pack(anchor='w')
        
        # 副标题
        subtitle = ttk.Label(header_frame, text="多公司混排 · 保留Excel格式 · 智能拆分", style='Subtitle.TLabel')
        subtitle.pack(anchor='w', pady=(5, 0))
        
        # 分隔线
        separator = tk.Frame(self.content_frame, height=2, bg=self.colors['border'])
        separator.pack(fill=tk.X, padx=30)
        
        # ========== 配置区域 ==========
        config_frame = tk.LabelFrame(self.content_frame, text=" 拆分设置 ", font=('Microsoft YaHei', 11, 'bold'),
                                   bg=self.colors['card'], fg=self.colors['text'], padx=15, pady=15)
        config_frame.pack(fill=tk.X, **padding)
        
        # Z012 + Z002
        self.create_field_with_hint(config_frame, "Z012 + Z002 每表行数", self.split_z012, "最少50行", width=10)
        
        # Z001
        self.create_field_with_hint(config_frame, "Z001 每表行数", self.split_z001, "最少50行", width=10)
        
        # ========== 数据文件选择区域（多文件） ==========
        data_frame = tk.LabelFrame(self.content_frame, text=" 数据文件选择 ", font=('Microsoft YaHei', 11, 'bold'),
                                   bg=self.colors['card'], fg=self.colors['text'], padx=15, pady=15)
        data_frame.pack(fill=tk.X, **padding)
        
        # 提示文字
        hint = tk.Label(data_frame, text="文件名格式: 公司代码-事业群.xlsx  (如: 4290-B.xlsx, 4291-A.xlsx)", 
                       font=('Microsoft YaHei', 9), bg=self.colors['card'], fg=self.colors['text_muted'])
        hint.pack(anchor='w', pady=(0, 10))
        
        # 已选文件列表框
        self.data_files_listbox = tk.Listbox(data_frame, height=6, font=('Consolas', 10),
                                             bg='white', fg=self.colors['text'],
                                             selectmode=tk.SINGLE, relief='solid', bd=1)
        self.data_files_listbox.pack(fill=tk.X, pady=(0, 10))
        
        # 按钮区域
        btn_frame = tk.Frame(data_frame, bg=self.colors['card'])
        btn_frame.pack(fill=tk.X)
        
        add_btn = tk.Button(btn_frame, text="+ 添加文件", command=self.add_data_files,
                           font=('Microsoft YaHei', 9), bg=self.colors['primary'], fg='white',
                           activebackground=self.colors['primary_dark'], bd=0, relief='flat',
                           padx=20, pady=6, cursor='hand2')
        add_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        clear_btn = tk.Button(btn_frame, text="清空全部", command=self.clear_data_files,
                             font=('Microsoft YaHei', 9), bg='white', fg=self.colors['text'],
                             activebackground=self.colors['bg'], bd=1, relief='solid',
                             padx=20, pady=6, cursor='hand2')
        clear_btn.pack(side=tk.LEFT)
        
        # ========== 其他文件选择区域 ==========
        file_frame = tk.LabelFrame(self.content_frame, text=" 其他文件 ", font=('Microsoft YaHei', 11, 'bold'),
                                   bg=self.colors['card'], fg=self.colors['text'], padx=15, pady=15)
        file_frame.pack(fill=tk.X, **padding)
        
        # 模板文件
        self.create_file_row(file_frame, "1. 模板文件", self.template_file,
                            "用于复制格式的空白模板文件", self.select_template_file)
        
        # 排除文件
        self.create_file_row(file_frame, "2. 排除文件", self.exclude_file,
                            "需要排除的客商编号列表（第一列）", self.select_exclude_file)
        
        # 输出目录
        self.create_file_row(file_frame, "输出目录", self.output_dir,
                            "处理结果保存位置", self.select_output_dir, is_dir=True)
        
        # ========== 日志区域 ==========
        log_frame = tk.LabelFrame(self.content_frame, text=" 处理日志 ", font=('Microsoft YaHei', 11, 'bold'),
                                  bg=self.colors['card'], fg=self.colors['text'], padx=15, pady=15)
        log_frame.pack(fill=tk.BOTH, expand=True, **padding)
        
        # 日志文本框
        self.log_text = tk.Text(log_frame, height=12, font=('Consolas', 10),
                                bg='#f1f5f9', fg=self.colors['text'],
                                relief='flat', padx=10, pady=10,
                                wrap=tk.WORD, state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # 滚动条
        log_scroll = ttk.Scrollbar(self.log_text, command=self.log_text.yview)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=log_scroll.set)
        
        # ========== 底部按钮区域 ==========
        bottom_frame = tk.Frame(self.content_frame, bg=self.colors['bg'])
        bottom_frame.pack(fill=tk.X, padx=30, pady=(0, 20))
        
        # 进度条（默认隐藏）
        self.progress = ttk.Progressbar(bottom_frame, mode='indeterminate', style='Horizontal.TProgressbar')
        self.progress.pack(fill=tk.X, pady=(0, 15))
        self.progress.pack_forget()
        
        # 按钮容器
        btn_frame = tk.Frame(bottom_frame, bg=self.colors['bg'])
        btn_frame.pack(fill=tk.X)
        
        # 清空日志按钮
        clear_log_btn = tk.Button(btn_frame, text="清空日志", command=self.clear_log,
                              font=('Microsoft YaHei', 10), bg='white', fg=self.colors['text_muted'],
                              activebackground=self.colors['bg'], activeforeground=self.colors['text'],
                              bd=1, relief='solid', padx=20, pady=8, cursor='hand2')
        clear_log_btn.pack(side=tk.LEFT)
        
        # 开始处理按钮
        self.start_btn = tk.Button(btn_frame, text="开始处理", command=self.start_process,
                                   font=('Microsoft YaHei', 12, 'bold'), bg=self.colors['primary'], fg='white',
                                   activebackground=self.colors['primary_dark'], activeforeground='white',
                                   bd=0, relief='flat', padx=40, pady=12, cursor='hand2')
        self.start_btn.pack(side=tk.RIGHT)
    
    def create_field_with_hint(self, parent, label_text, variable, hint_text, width=20):
        """创建带提示的输入字段"""
        frame = tk.Frame(parent, bg=self.colors['card'])
        frame.pack(fill=tk.X, pady=8)
        
        label_frame = tk.Frame(frame, bg=self.colors['card'])
        label_frame.pack(fill=tk.X)
        
        label = tk.Label(label_frame, text=label_text, font=('Microsoft YaHei', 10, 'bold'),
                         bg=self.colors['card'], fg=self.colors['text'])
        label.pack(side=tk.LEFT)
        
        hint = tk.Label(label_frame, text=hint_text, font=('Microsoft YaHei', 9),
                        bg=self.colors['card'], fg=self.colors['text_muted'])
        hint.pack(side=tk.LEFT, padx=(10, 0))
        
        entry = tk.Entry(frame, textvariable=variable, font=('Microsoft YaHei', 11),
                         width=width, relief='solid', bd=1, bg='white', justify='center')
        entry.pack(fill=tk.X, pady=(5, 0), ipady=6)
    
    def create_file_row(self, parent, label_text, variable, hint_text, command, is_dir=False):
        """创建文件选择行"""
        frame = tk.Frame(parent, bg=self.colors['card'])
        frame.pack(fill=tk.X, pady=10)
        
        # 标签
        label = tk.Label(frame, text=label_text, font=('Microsoft YaHei', 10, 'bold'),
                         bg=self.colors['card'], fg=self.colors['text'])
        label.pack(anchor='w')
        
        # 提示文字
        hint = tk.Label(frame, text=hint_text, font=('Microsoft YaHei', 9),
                        bg=self.colors['card'], fg=self.colors['text_muted'])
        hint.pack(anchor='w', pady=(2, 5))
        
        # 输入框和按钮容器
        input_frame = tk.Frame(frame, bg=self.colors['card'])
        input_frame.pack(fill=tk.X)
        
        entry = tk.Entry(input_frame, textvariable=variable, font=('Microsoft YaHei', 10),
                         relief='solid', bd=1, bg='white')
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=6)
        
        btn = tk.Button(input_frame, text="浏览...", command=lambda: command(is_dir) if is_dir else command(),
                        font=('Microsoft YaHei', 9), bg='white', fg=self.colors['text'],
                        activebackground=self.colors['bg'], bd=1, relief='solid',
                        padx=15, pady=4, cursor='hand2')
        btn.pack(side=tk.RIGHT, padx=(10, 0))
    
    # ========== 事件处理 ==========
    def on_frame_configure(self, event=None):
        """框架大小改变时更新Canvas滚动区域"""
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
    
    def on_canvas_configure(self, event):
        """Canvas大小改变时调整内容宽度"""
        self.main_canvas.itemconfig(self.canvas_window, width=event.width)
    
    def on_mousewheel(self, event):
        """鼠标滚轮滚动"""
        self.main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    # ========== 多文件管理 ==========
    def parse_filename(self, filepath):
        """从文件名解析公司代码和事业群"""
        filename = os.path.basename(filepath)
        name_without_ext = os.path.splitext(filename)[0]
        
        # 匹配格式: 公司代码-事业群 (如 4290-B, 4291-A)
        # 公司代码: 数字
        # 事业群: 字母或数字
        match = re.match(r'^(\d+)-([A-Za-z0-9]+)$', name_without_ext)
        
        if match:
            company_code = match.group(1)
            business_group = match.group(2).upper()
            return company_code, business_group
        else:
            return None, None
    
    def add_data_files(self):
        """添加数据文件（多选）"""
        files = filedialog.askopenfilenames(
            title="选择数据文件（可多选）",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        
        if not files:
            return
        
        errors = []
        for filepath in files:
            # 检查是否已存在
            if any(f['path'] == filepath for f in self.data_files):
                continue
            
            # 解析文件名
            company_code, business_group = self.parse_filename(filepath)
            
            if company_code is None:
                errors.append(f"{os.path.basename(filepath)}: 文件名格式错误，应为 '公司代码-事业群.xlsx'")
                continue
            
            # 添加到列表
            self.data_files.append({
                'path': filepath,
                'company': company_code,
                'group': business_group
            })
        
        # 更新显示
        self.update_data_files_list()
        
        # 报告错误
        if errors:
            messagebox.showwarning("文件名格式警告", "以下文件未添加:\n\n" + "\n".join(errors))
    
    def update_data_files_list(self):
        """更新文件列表显示"""
        self.data_files_listbox.delete(0, tk.END)
        for f in self.data_files:
            display = f"{f['company']}-{f['group']}: {os.path.basename(f['path'])}"
            self.data_files_listbox.insert(tk.END, display)
    
    def clear_data_files(self):
        """清空所有数据文件"""
        self.data_files = []
        self.data_files_listbox.delete(0, tk.END)
    
    # ========== 文件选择 ==========
    def select_template_file(self):
        file = filedialog.askopenfilename(
            title="选择模板文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file:
            self.template_file.set(file)
    
    def select_exclude_file(self):
        file = filedialog.askopenfilename(
            title="选择排除文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file:
            self.exclude_file.set(file)
    
    def select_output_dir(self, is_dir=True):
        dir = filedialog.askdirectory(title="选择输出目录")
        if dir:
            self.output_dir.set(dir)
    
    # ========== 日志功能 ==========
    def log(self, msg):
        """添加日志"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.root.update()
    
    def clear_log(self):
        """清空日志"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)
    
    # ========== 处理逻辑 ==========
    def start_process(self):
        """开始处理"""
        # 验证输入
        if len(self.data_files) == 0:
            messagebox.showerror("错误", "请至少选择一个数据文件")
            return
        
        template_file = self.template_file.get().strip()
        if not template_file or not os.path.exists(template_file):
            messagebox.showerror("错误", "请选择有效的模板文件")
            return
        
        exclude_file = self.exclude_file.get().strip()
        if not exclude_file or not os.path.exists(exclude_file):
            messagebox.showerror("错误", "请选择有效的排除文件")
            return
        
        output_dir = self.output_dir.get().strip()
        if not output_dir:
            output_dir = os.path.expanduser("~/Desktop")
        
        try:
            split_z012 = int(self.split_z012.get() or 200)
            split_z001 = int(self.split_z001.get() or 500)
        except ValueError:
            messagebox.showerror("错误", "拆分行数必须是数字")
            return
        
        # 更新UI状态
        self.start_btn.config(state=tk.DISABLED, text="处理中...", bg=self.colors['text_muted'])
        self.progress.pack(fill=tk.X, pady=(0, 15))
        self.progress.start()
        self.clear_log()
        
        try:
            self.process_excel_multi(self.data_files, template_file, exclude_file, output_dir, 
                                    split_z012, split_z001)
        except Exception as e:
            self.log(f"\n❌ 错误: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            messagebox.showerror("错误", str(e))
        finally:
            self.progress.stop()
            self.progress.pack_forget()
            self.start_btn.config(state=tk.NORMAL, text="开始处理", bg=self.colors['primary'])
    
    def process_excel_multi(self, data_files, template_file, exclude_file, output_dir, 
                           split_z012_z002=200, split_z001=500):
        """处理多公司Excel - 混排模式"""
        
        self.log(f"🚀 开始处理多公司混排模式")
        self.log(f"📋 共 {len(data_files)} 个公司数据文件")
        self.log(f"📋 拆分规则: Z012+Z002={split_z012_z002}行/表, Z001={split_z001}行/表")
        
        # 显示各公司信息
        for f in data_files:
            self.log(f"   • {f['company']}-{f['group']}: {os.path.basename(f['path'])}")
        
        # 清理输出目录
        output_path = os.path.join(output_dir, "多公司混排输出")
        if os.path.exists(output_path):
            shutil.rmtree(output_path)
        os.makedirs(output_path, exist_ok=True)
        
        # 读取排除文件（共用）
        self.log(f"\n📂 读取排除文件...")
        wb_exclude = load_workbook(exclude_file)
        exclude_ws = wb_exclude[wb_exclude.sheetnames[0]]
        self.log(f"   排除表单: {wb_exclude.sheetnames[0]}")
        
        # 1. 读取排除列表
        self.log(f"\n🔍 读取排除列表...")
        exclude_codes = set()
        for row in range(2, exclude_ws.max_row + 1):
            code = exclude_ws.cell(row=row, column=1).value
            if code:
                code_str = str(code).strip()
                exclude_codes.add(code_str)
                exclude_codes.add(code_str.lstrip('0') or '0')
        self.log(f"   排除客商: {len(exclude_codes)//2} 条")
        
        # 2. 处理每个公司的数据文件
        all_data = []  # 所有公司的数据合并
        total_manual = {}  # 需手工处理的客商（跨公司去重）
        
        for file_info in data_files:
            company_code = file_info['company']
            business_group = file_info['group']
            data_file = file_info['path']
            
            self.log(f"\n📊 处理公司 {company_code}-{business_group}...")
            
            # 读取数据文件
            wb_data = load_workbook(data_file)
            
            # 按名称找客户和供应商表单
            customer_ws = None
            supplier_ws = None
            
            for name in wb_data.sheetnames:
                lower = name.lower()
                if "客户" in lower:
                    customer_ws = wb_data[name]
                elif "供应" in lower:
                    supplier_ws = wb_data[name]
            
            if not customer_ws or not supplier_ws:
                raise Exception(f"{os.path.basename(data_file)} 缺少客户或供应商表单！")
            
            self.log(f"   客户表单: {customer_ws.title}, 供应商表单: {supplier_ws.title}")
            
            # 处理客户
            customer_count = 0
            customer_codes_seen = set()
            
            for row in range(2, customer_ws.max_row + 1):
                code = customer_ws.cell(row=row, column=6).value
                if code:
                    code_str = str(code).strip()
                    code_no_zero = code_str.lstrip('0') or '0'
                    if code_str not in customer_codes_seen:
                        customer_codes_seen.add(code_str)
                        group = str(customer_ws.cell(row=row, column=4).value or 'Z001').strip()
                        
                        if code_str in exclude_codes or code_no_zero in exclude_codes:
                            # 需手工处理 - 用公司代码+客商编号作为key去重
                            key = f"{company_code}_{code_str}"
                            if key not in total_manual:
                                total_manual[key] = {
                                    'code': code_str, 
                                    'group': group,
                                    'company': company_code,
                                    'business_group': business_group
                                }
                        else:
                            all_data.append({
                                'code': code_str, 
                                'group': group, 
                                'type': 'FLCU00',
                                'company': company_code,
                                'business_group': business_group
                            })
                            customer_count += 1
            
            # 处理供应商
            supplier_count = 0
            supplier_codes_seen = set()
            
            for row in range(2, supplier_ws.max_row + 1):
                code = supplier_ws.cell(row=row, column=6).value
                if code:
                    code_str = str(code).strip()
                    code_no_zero = code_str.lstrip('0') or '0'
                    if code_str not in supplier_codes_seen:
                        supplier_codes_seen.add(code_str)
                        group = str(supplier_ws.cell(row=row, column=4).value or 'Z001').strip()
                        
                        if code_str in exclude_codes or code_no_zero in exclude_codes:
                            key = f"{company_code}_{code_str}"
                            if key not in total_manual:
                                total_manual[key] = {
                                    'code': code_str, 
                                    'group': group,
                                    'company': company_code,
                                    'business_group': business_group
                                }
                        else:
                            all_data.append({
                                'code': code_str, 
                                'group': group, 
                                'type': 'FLVN00',
                                'company': company_code,
                                'business_group': business_group
                            })
                            supplier_count += 1
            
            self.log(f"   客户: {len(customer_codes_seen)} 条 → 有效 {customer_count}")
            self.log(f"   供应商: {len(supplier_codes_seen)} 条 → 有效 {supplier_count}")
        
        # 3. 合并排序（混排模式）
        self.log(f"\n🔄 合并并排序（混排模式）...")
        group_weight = {'Z012': 3, 'Z002': 2, 'Z001': 1}
        
        # 排序优先级: 分组 > 客商编号 > 公司代码
        all_data.sort(key=lambda x: (
            -group_weight.get(x['group'], 0),  # 分组优先级（Z012>Z002>Z001）
            x['code'],                          # 客商编号升序（同客商挨着）
            x['company']                        # 公司代码升序
        ))
        
        z012_z002_data = [d for d in all_data if d['group'] in ['Z012', 'Z002']]
        z001_data = [d for d in all_data if d['group'] == 'Z001']
        manual_codes = list(total_manual.values())
        
        self.log(f"   总计: {len(all_data)} 条")
        self.log(f"   ├─ Z012+Z002: {len(z012_z002_data)} 条")
        self.log(f"   ├─ Z001: {len(z001_data)} 条")
        self.log(f"   └─ 需手工处理: {len(manual_codes)} 条")
        
        # 4. 生成完整模板
        self.log(f"\n💾 生成完整模板...")
        wb_full = load_workbook(template_file)
        ws_full = wb_full[wb_full.sheetnames[0]]
        
        for idx, item in enumerate(all_data, start=3):
            ws_full.cell(row=idx, column=1, value=item['business_group'])  # 事业群
            ws_full.cell(row=idx, column=2, value=item['code'])            # 客商编号
            ws_full.cell(row=idx, column=3, value=item['type'])            # 客商类型
            ws_full.cell(row=idx, column=4, value=item['company'])         # 公司代码
            if ws_full.max_column >= 6:
                ws_full.cell(row=idx, column=6, value=item['group'])       # 分组
        
        wb_full.save(os.path.join(output_path, '完整模板.xlsx'))
        self.log(f"   ✓ 完整模板: {len(all_data)} 行")
        
        # 5. 生成手工处理文件（按公司分表）
        if manual_codes:
            self.log(f"\n💾 生成手工处理清单（按公司分表）...")
            
            # 按公司分组
            manual_by_company = {}
            for item in manual_codes:
                company = item['company']
                if company not in manual_by_company:
                    manual_by_company[company] = []
                manual_by_company[company].append(item)
            
            # 每个公司一个文件
            for company, items in sorted(manual_by_company.items()):
                wb_manual = openpyxl.Workbook()
                ws_manual = wb_manual.active
                ws_manual.title = '手工处理客商'
                ws_manual.cell(row=1, column=1, value='公司代码')
                ws_manual.cell(row=1, column=2, value='事业群')
                ws_manual.cell(row=1, column=3, value='客商编号')
                ws_manual.cell(row=1, column=4, value='分组')
                for idx, item in enumerate(items, start=2):
                    ws_manual.cell(row=idx, column=1, value=item['company'])
                    ws_manual.cell(row=idx, column=2, value=item['business_group'])
                    ws_manual.cell(row=idx, column=3, value=item['code'])
                    ws_manual.cell(row=idx, column=4, value=item['group'])
                wb_manual.save(os.path.join(output_path, f'{company}需要手工处理的客商.xlsx'))
                self.log(f"   ✓ {company}: {len(items)} 条")
        
        # 6. 拆分小表
        self.log(f"\n✂️ 拆分小表...")
        
        def split_data(data, chunk_size):
            chunks = []
            for i in range(0, len(data), chunk_size):
                chunks.append(data[i:i+chunk_size])
            if len(chunks) > 1 and len(chunks[-1]) < 50:
                tail = chunks.pop()
                chunks[-1].extend(tail)
                self.log(f"   📝 小尾巴合并: {len(tail)} 行 → 并入前表")
            return chunks
        
        z012_z002_chunks = split_data(z012_z002_data, split_z012_z002)
        z001_chunks = split_data(z001_data, split_z001)
        all_chunks = z012_z002_chunks + z001_chunks
        
        self.log(f"   Z012+Z002: {len(z012_z002_data)} 条 → {len(z012_z002_chunks)} 个文件")
        self.log(f"   Z001: {len(z001_data)} 条 → {len(z001_chunks)} 个文件")
        
        # 7. 生成拆分文件
        self.log(f"\n💾 生成拆分文件...")
        for idx, chunk in enumerate(all_chunks, 1):
            wb_chunk = load_workbook(template_file)
            ws_chunk = wb_chunk[wb_chunk.sheetnames[0]]
            
            # 填充数据
            for row_idx, item in enumerate(chunk, start=3):
                ws_chunk.cell(row=row_idx, column=1, value=item['business_group'])
                ws_chunk.cell(row=row_idx, column=2, value=item['code'])
                ws_chunk.cell(row=row_idx, column=3, value=item['type'])
                ws_chunk.cell(row=row_idx, column=4, value=item['company'])
            
            # 删除数据行之后的空行（防止SAP报错）
            last_data_row = 2 + len(chunk)  # 表头2行 + 数据行
            max_row = ws_chunk.max_row
            if max_row > last_data_row:
                # 删除多余的行
                ws_chunk.delete_rows(last_data_row + 1, max_row - last_data_row)
            
            wb_chunk.save(os.path.join(output_path, f'{idx:02d}.xlsx'))
        
        self.log(f"\n" + "="*50)
        self.log(f"✅ 处理完成！")
        self.log(f"📁 输出目录: {output_path}")
        self.log(f"📊 统计: 总数据 {len(all_data)} | 手工处理 {len(manual_codes)} | 拆分文件 {len(all_chunks)}")
        self.log(f"="*50)
        
        messagebox.showinfo("处理完成", 
            f"✅ 处理成功！\n\n"
            f"📊 数据统计:\n"
            f"   公司数量: {len(data_files)} 家\n"
            f"   总数据: {len(all_data)} 条\n"
            f"   需手工处理: {len(manual_codes)} 条\n"
            f"   拆分文件: {len(all_chunks)} 个\n\n"
            f"📁 输出目录:\n{output_path}")


def main():
    root = tk.Tk()
    app = ModernSAPMDGTool(root)
    root.mainloop()


if __name__ == "__main__":
    main()
